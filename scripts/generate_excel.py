"""
Recibe JSON por stdin con { template_path, domicilio: [...], sucursal: [...] }
Escribe los datos en la plantilla Andreani usando openpyxl (mismo método que web_app original)
Devuelve el archivo .xlsx codificado en base64 por stdout
"""
import sys
import json
import base64
import os
import tempfile
from openpyxl import load_workbook

FILA_INICIO = 3
ULTIMA_FILA = 400


def limpiar_numero_calle(valor):
    """Igual que web_app original: S/N, -, vacío → "0". Solo dígitos."""
    if valor is None:
        return "0"
    s = str(valor).strip().upper()
    if s in ("SN", "S/N", "S N", "-", ""):
        return "0"
    if s.endswith(".0"):
        s = s[:-2]
    dig = "".join(ch for ch in s if ch.isdigit())
    return dig if dig else "0"


def limpiar_piso(valor):
    """
    Piso: quita caracteres inválidos para Andreani.
    - Si el campo está vacío, es '-', 'S/N', '*' etc → vacío
    - Trunca en el primer separador seguido de texto libre
      (ej: "2do B - Porteria" → "2do B", "PB/Planta Baja" → "PB")
    - Solo conserva letras, dígitos y espacios (Andreani rechaza *, #, /, etc.)
    """
    if valor is None:
        return ""
    s = str(valor).strip()
    if s in ("-", "S/N", "s/n", "SN", "sn", "0", "", "*"):
        return ""
    # Truncar en el primer '-' o '/' que actúe como separador de texto libre
    import re
    match = re.search(r'\s*[-/]\s*[a-zA-Z]', s)
    if match:
        s = s[:match.start()].strip()
    # Conservar solo letras, dígitos y espacios (elimina *, #, -, /, etc.)
    s = re.sub(r'[^a-zA-Z0-9áéíóúÁÉÍÓÚüÜñÑ ]', '', s).strip()
    return s


def limpiar_telefono(raw):
    """Devuelve solo dígitos como entero. Remueve código 54 del inicio si está."""
    if not raw:
        return None
    digits = "".join(c for c in str(raw) if c.isdigit())
    if not digits:
        return None
    # Remueve el código de país 54 si viene incluido
    if digits.startswith("54") and len(digits) > 10:
        digits = digits[2:]
    if digits.startswith("0") and len(digits) > 8:
        digits = digits[1:]
    try:
        return int(digits)
    except ValueError:
        return None


def generate_excel(data):
    template_path = data["template_path"]
    domicilio     = data.get("domicilio", [])
    sucursal      = data.get("sucursal", [])

    wb     = load_workbook(template_path, data_only=False)
    ws_dom = wb["A domicilio"]
    ws_suc = wb["A sucursal"]

    # --- Limpiar filas de datos (preserva encabezados y Configuracion) ---
    for ws in [ws_dom, ws_suc]:
        for row in range(FILA_INICIO, ULTIMA_FILA + 1):
            for col in range(1, 20):
                ws.cell(row=row, column=col).value = None

    # --- Escribir hoja "A domicilio" ---
    # Mapeo columna → campo (igual que web_app original, 1-based)
    # Col 1:  Paquete Guardado  (vacío)
    # Col 2:  Peso (grs)
    # Col 3:  Alto (cm)
    # Col 4:  Ancho (cm)
    # Col 5:  Profundidad (cm)
    # Col 6:  Valor declarado
    # Col 7:  Numero Interno
    # Col 8:  Nombre
    # Col 9:  Apellido
    # Col 10: DNI
    # Col 11: Email
    # Col 12: Celular código
    # Col 13: Celular número  ← ENTERO
    # Col 14: Calle
    # Col 15: Número
    # Col 16: Piso
    # Col 17: Departamento
    # Col 18: Provincia / Localidad / CP  ← debe coincidir con lista desplegable
    # Col 19: Observaciones
    for i, item in enumerate(domicilio):
        row = FILA_INICIO + i
        ws_dom.cell(row=row, column=2).value  = item.get("Peso (grs)", 1)
        ws_dom.cell(row=row, column=3).value  = item.get("Alto (cm)", 1)
        ws_dom.cell(row=row, column=4).value  = item.get("Ancho (cm)", 1)
        ws_dom.cell(row=row, column=5).value  = item.get("Profundidad (cm)", 1)
        ws_dom.cell(row=row, column=6).value  = item.get("Valor declarado ($ c/IVA)", 6000)
        ws_dom.cell(row=row, column=7).value  = str(item.get("Numero Interno", ""))
        ws_dom.cell(row=row, column=8).value  = item.get("Nombre", "")
        ws_dom.cell(row=row, column=9).value  = item.get("Apellido", "")
        ws_dom.cell(row=row, column=10).value = str(item.get("DNI", ""))
        ws_dom.cell(row=row, column=11).value = item.get("Email", "")
        ws_dom.cell(row=row, column=12).value = str(item.get("Celular código", "54"))
        ws_dom.cell(row=row, column=13).value = limpiar_telefono(item.get("Celular número", ""))
        ws_dom.cell(row=row, column=14).value = item.get("Calle", "")
        ws_dom.cell(row=row, column=15).value = limpiar_numero_calle(item.get("Número", ""))
        ws_dom.cell(row=row, column=16).value = limpiar_piso(item.get("Piso", ""))
        ws_dom.cell(row=row, column=17).value = ""
        ws_dom.cell(row=row, column=18).value = item.get("Provincia / Localidad / CP", "")
        ws_dom.cell(row=row, column=19).value = item.get("Observaciones", "")

    # --- Escribir hoja "A sucursal" ---
    # Col 1-13: igual que domicilio
    # Col 14: Sucursal  ← debe coincidir con lista desplegable
    for i, item in enumerate(sucursal):
        row = FILA_INICIO + i
        ws_suc.cell(row=row, column=2).value  = item.get("Peso (grs)", 1)
        ws_suc.cell(row=row, column=3).value  = item.get("Alto (cm)", 1)
        ws_suc.cell(row=row, column=4).value  = item.get("Ancho (cm)", 1)
        ws_suc.cell(row=row, column=5).value  = item.get("Profundidad (cm)", 1)
        ws_suc.cell(row=row, column=6).value  = item.get("Valor declarado ($ c/IVA)", 6000)
        ws_suc.cell(row=row, column=7).value  = str(item.get("Numero Interno", ""))
        ws_suc.cell(row=row, column=8).value  = item.get("Nombre", "")
        ws_suc.cell(row=row, column=9).value  = item.get("Apellido", "")
        ws_suc.cell(row=row, column=10).value = str(item.get("DNI", ""))
        ws_suc.cell(row=row, column=11).value = item.get("Email", "")
        ws_suc.cell(row=row, column=12).value = str(item.get("Celular código", "54"))
        ws_suc.cell(row=row, column=13).value = limpiar_telefono(item.get("Celular número", ""))
        ws_suc.cell(row=row, column=14).value = item.get("Sucursal", "")

    # --- Guardar en archivo temporal y devolver como base64 ---
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path = f.name

    wb.save(tmp_path)

    with open(tmp_path, "rb") as f:
        content = f.read()

    os.unlink(tmp_path)
    return base64.b64encode(content).decode("utf-8")


if __name__ == "__main__":
    data   = json.loads(sys.stdin.buffer.read().decode("utf-8"))
    result = generate_excel(data)
    sys.stdout.write(result)
